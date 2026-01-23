# utils/excel.py
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from config import EXCEL_PATH, LOG_DIR

# ==================================================
# Filename parser
#   5G_20MHz_STA_RX_CH36_MCS8.txt
# ==================================================

FILENAME_RE = re.compile(
    r"""
    ^5G_
    (?P<bw>\d+)MHz_
    (?P<role>AP|STA)_
    (?P<direction>TX|RX)_
    CH(?P<ch>\d+)_
    MCS(?P<mcs>\d+)
    (?:_.*)?\.txt$
    """,
    re.VERBOSE,
)

# ==================================================
# Bitrate extractor
# ==================================================

BITRATE_RE = re.compile(r"(?P<rate>\d+(?:\.\d+)?)\s+Mbits/sec")

# ==================================================
# ===== Stability Matrix (RAW DATA) =====
# value = (sheet_name, start_col)
# ==================================================

STABILITY_MAP = {
    # ===== STA_TX =====
    ("STA", "TX", 20, 36): ("5GHz_Tput Stability_20MHz_STA", 2),
    ("STA", "TX", 20, 149): ("5GHz_Tput Stability_20MHz_STA", 44),
    ("STA", "TX", 40, 36): ("5GHz_Tput Stability_40MHz_STA", 2),
    ("STA", "TX", 40, 149): ("5GHz_Tput Stability_40MHz_STA", 46),
    ("STA", "TX", 80, 36): ("5GHz_Tput Stability_80MHz_STA", 2),
    ("STA", "TX", 80, 149): ("5GHz_Tput Stability_80MHz_STA", 46),

    # ===== AP_TX =====
    ("AP", "TX", 20, 36): ("5GHz_Tput Stability_20MHz_AP", 2),
    ("AP", "TX", 20, 149): ("5GHz_Tput Stability_20MHz_AP", 44),
    ("AP", "TX", 40, 36): ("5GHz_Tput Stability_40MHz_AP", 2),
    ("AP", "TX", 40, 149): ("5GHz_Tput Stability_40MHz_AP", 46),
    ("AP", "TX", 80, 36): ("5GHz_Tput Stability_80MHz_AP", 2),
    ("AP", "TX", 80, 149): ("5GHz_Tput Stability_80MHz_AP", 46),

    # ===== STA_RX =====
    ("STA", "RX", 20, 36): ("5GHz_Tput Stability_20MHz_STA", 23),
    ("STA", "RX", 20, 149): ("5GHz_Tput Stability_20MHz_STA", 65),
    ("STA", "RX", 40, 36): ("5GHz_Tput Stability_40MHz_STA", 24),
    ("STA", "RX", 40, 149): ("5GHz_Tput Stability_40MHz_STA", 68),
    ("STA", "RX", 80, 36): ("5GHz_Tput Stability_80MHz_STA", 24),
    ("STA", "RX", 80, 149): ("5GHz_Tput Stability_80MHz_STA", 68),

    # ===== AP_RX =====
    ("AP", "RX", 20, 36): ("5GHz_Tput Stability_20MHz_AP", 23),
    ("AP", "RX", 20, 149): ("5GHz_Tput Stability_20MHz_AP", 65),
    ("AP", "RX", 40, 36): ("5GHz_Tput Stability_40MHz_AP", 24),
    ("AP", "RX", 40, 149): ("5GHz_Tput Stability_40MHz_AP", 68),
    ("AP", "RX", 80, 36): ("5GHz_Tput Stability_80MHz_AP", 24),
    ("AP", "RX", 80, 149): ("5GHz_Tput Stability_80MHz_AP", 68),
}

# ==================================================
# ===== Average Matrix =====
# value = (row, start_col)
# ==================================================

AVERAGE_MAP = {
    # ===== STA_TX =====
    ("STA", "TX", 20, 36): (8, 8),
    ("STA", "TX", 20, 149): (16, 8),
    ("STA", "TX", 40, 36): (6, 7),
    ("STA", "TX", 40, 149): (14, 7),
    ("STA", "TX", 80, 36): (4, 7),
    ("STA", "TX", 80, 149): (12, 7),

    # ===== AP_TX =====
    ("AP", "TX", 20, 36): (43, 8),
    ("AP", "TX", 20, 149): (51, 8),
    ("AP", "TX", 40, 36): (41, 7),
    ("AP", "TX", 40, 149): (49, 7),
    ("AP", "TX", 80, 36): (39, 7),
    ("AP", "TX", 80, 149): (47, 7),

    # ===== STA_RX =====
    ("STA", "RX", 20, 36): (25, 8),
    ("STA", "RX", 20, 149): (33, 8),
    ("STA", "RX", 40, 36): (23, 7),
    ("STA", "RX", 40, 149): (31, 7),
    ("STA", "RX", 80, 36): (21, 7),
    ("STA", "RX", 80, 149): (29, 7),

    # ===== AP_RX =====
    ("AP", "RX", 20, 36): (60, 8),
    ("AP", "RX", 20, 149): (68, 8),
    ("AP", "RX", 40, 36): (58, 7),
    ("AP", "RX", 40, 149): (66, 7),
    ("AP", "RX", 80, 36): (56, 7),
    ("AP", "RX", 80, 149): (64, 7),
}

# ==================================================
# Log parsing
# ==================================================

def parse_log(path: Path):
    interval_rates = []
    sender_avg = None

    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()

            if "sender" in line and "Mbits/sec" in line:
                m = BITRATE_RE.search(line)
                if m:
                    sender_avg = float(m.group("rate"))
                continue

            if "Mbits/sec" in line and "-" in line and "receiver" not in line:
                m = BITRATE_RE.search(line)
                if m:
                    interval_rates.append(float(m.group("rate")))

    if not interval_rates or sender_avg is None:
        return None, None

    return interval_rates[:30], sender_avg

# ==================================================
# Main
# ==================================================

def main():
    log_root = Path(LOG_DIR)
    logs = sorted(log_root.rglob("*.txt"))

    if not logs:
        print("❌ No log files found")
        return

    wb = load_workbook(EXCEL_PATH)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = EXCEL_PATH.replace(".xlsx", f"_filled_{ts}.xlsx")

    for log in logs:
        m = FILENAME_RE.search(log.name)
        if not m:
            print(f"[SKIP] filename not matched: {log.name}")
            continue

        bw = int(m.group("bw"))
        role = m.group("role")
        direction = m.group("direction")
        ch = int(m.group("ch"))
        mcs = int(m.group("mcs"))

        key = (role, direction, bw, ch)

        interval_rates, sender_avg = parse_log(log)
        if interval_rates is None:
            print(f"[SKIP] invalid content: {log.name}")
            continue

        # -------- Stability --------
        if key not in STABILITY_MAP:
            print(f"[SKIP] no stability mapping: {key}")
            continue

        sheet_name, start_col = STABILITY_MAP[key]
        if sheet_name not in wb.sheetnames:
            print(f"[WARN] missing sheet: {sheet_name}")
            continue

        ws_s = wb[sheet_name]
        max_mcs = 8 if bw == 20 else 9
        col = start_col + (max_mcs - mcs)

        for i, rate in enumerate(interval_rates):
            ws_s.cell(row=3 + i, column=col, value=rate)

        print(f"[WRITE] {key} → {sheet_name} col {col}")

        # -------- Average --------
        if key in AVERAGE_MAP:
            ws_a = wb["5GHz_Average Tput per Rate"]
            row_a, start_col_a = AVERAGE_MAP[key]
            col_a = start_col_a + (max_mcs - mcs)
            ws_a.cell(row=row_a, column=col_a, value=sender_avg)
            print(f"[WRITE] Average row {row_a}, col {col_a}")

    wb.save(out_path)
    print(f"\n✅ Excel written: {out_path}")


if __name__ == "__main__":
    main()
